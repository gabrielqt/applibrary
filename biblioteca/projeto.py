#tratar erros do usuario, fazer função daquele for value, desenvolver pós tela de login.

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import json
from time import sleep
import os
import pandas as pd
from tabulate import tabulate

ROOT = Path(__file__).parent
BOOK_PATH = ROOT / 'books.xlsx'
USERS_PATH = ROOT / 'users.json'

genr = 100
favgen = list()
menu = ''
userdata = {}
login_ = False
password_ = False
newbook = list()
library = pd.read_excel(BOOK_PATH)
oi = 1

def clear():
    os.system('cls')
clear()


#verifica se o arquivo de planilha já existe
if BOOK_PATH.exists():
    workbook = load_workbook(BOOK_PATH)

worksheet: Worksheet = workbook.active

def wait():
    for _ in range(6):
        print('...', end='', flush=True)
        sleep(0.2)
        print('\b\b\b   \b\b\b', end='', flush=True)
        sleep(0.2)



#library:

def showlib ():
    print(tabulate(library, headers = 'keys', tablefmt='pretty'))





#USERS:
if USERS_PATH.exists():
    with open(USERS_PATH, 'r+', encoding='utf8') as file:      #se existir vai retorna dict a variavel userdata
        userdata = json.load(file)

        
def save_userdata(userdata):
    with open (USERS_PATH, 'w', encoding = 'utf8') as file:  #atualizando json/salvando usuario
        json.dump(userdata,file,indent=2)
        
def genid ():         #gerar uma id
    if userdata == {}:
        return 1
    else:
        userid = int(max(userdata.keys())) + 1
        return userid
    
    
    
    
def genres():
    global genr
    match genr:
        case 1:
            favgen.append('Romance')
        case 2:
            favgen.append('Fantasia')
        case 3:
            favgen.append('Ficcao')
        case 4:
            favgen.append('Realismo Mágico') 
        case 5:
            favgen.append('Fábula')
        case 6:
            favgen.append('Aventura')                                      
        case 7:
            favgen.append('Mistério')
        case 8:
            favgen.append('Terror')
        case 9:
            favgen.append('Romance')
        
        
        
        
        
            
def create_user(login,password,name,age,favgen):
    userid = genid()
    userdata[userid] = {'login': login, 'password': password, 'name': name, 'age': age, 'favgen': favgen, 'books' : [] }
    save_userdata(userdata)

        
    
#START:

while True:
    print('-=-'*20)
    print('''                       \033[33mBIBLIOTECA ONLINE\033[m''')
    print('-=-'*20)

    print('1 - Já tenho uma conta (Fazer login)')
    print('2 - Não tenho uma conta (Criar conta)')
    print()
    menu = str(input('Digite a opção >> '))
    
    
    
    
    if menu == '1':
        login_ = str(input('LOGIN: '))    
        passw_ = str(input('PASSWORD: '))
        currentuser = None
        for user_id, user_data in userdata.items():

            if login_ == user_data['login']:
                currentid = user_id
                currentuser = user_data
                login_ = True
            if currentuser and currentuser['password'] == passw_:
                passw_ = True
        
        if not (login_ and passw_):
            print()
            print('Login ou senha incorreto!')
            sleep(2)
            clear()
        else:
            name_ = ''.join(currentuser['name'])
            print()
            print(f'Seja Bem-Vindo {name_}!')
            wait()
            sleep(1)
            break
                    
 
                        
                    
    
    
    elif menu == '2':
        while True:
            login = str(input('Digite o login de sua preferência: ')).strip()
            if len(login) < 8:
                print('Seu login deve ter no minímo 8 caracteres!')
            else:
                break
        for value in userdata.values():
            for item in value:
                if value[item] == login:
                    print('Esse usuário de login já existe!')
                    break
        password = str(input('Digite uma senha: ')).strip()
        password2 = str(input('Repita sua senha: '))
        while password2 != password:
            print('As senhas não conferem, digite uma nova senha novamente!')
            sleep (0.4)
            password = str(input('Digite uma senha: '))
            password2 = str(input('Repita sua senha: '))
        name = str(input('Digite seu nome completo: ')).strip().title()
        age = int(input('Digite sua idade: '))
        print('''
GENÊROS TEXTUAIS:                       
    1.  Romance
    2.  Fantasia
    3.  Ficção
    4.  Realismo Mágico
    5.  Fábula
    6.  Aventura
    7.  Mistério
    8.  Terror
    9.  Romance
    10. SAIR
''')
        while genr is not 10:
            genr = int(input('Digite uma opção: '))
            genres()
            
        create_user(login,password,name,age,favgen)
        
        print('Criando sua conta', end = '')
        wait()
        print('\nConta criada com sucesso!')
        sleep(1.0)
        clear()
            
            

        
    else:
        clear()
        print('Opção inválida :( Digite novamente.')
        
#MENU



while True:
    print('_'*45)
    print('            \033[35mMenu de Usuário\033[m')
    print('_'*45)
    print()
    print(f'\tUser:{currentuser['login']} - ID: {currentid}')
    print(f'\t{currentuser['name']},', end=' ')
    print(f'{currentuser['age']}')
    print()
    print('1 - Mostrar biblioteca.')
    print('2 - Adicionar livro a biblioteca.')
    print('3 - Remover livro da biblioteca.')
    print('4 - SAIR.')
    
    requser = str(input('Digite a opção: '))

    match requser:
        
        case '1':
            showlib()
            
        case '2':
            print('Vamos adicionar um novo livro a nossa base de dados!')
            sleep(1)
            print()
            title = str(input('Digite o títutlo dele: ')).title()
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == title:
                        print('Este livro já está em nossa biblioteca!')
            if 'python' in title:
                print('Ótima escolha!')
            newbook.append(title)
            author = str(input('Digite o autor do livro: '))
            newbook.append(author)
            genre = str(input('Digite o seu genêro: '))
            newbook.append(genre)
            year = str(input(f'Digite o ano de publicação do {title}: '))
            newbook.append(year)
            worksheet.append(newbook)
            print('Adicionando livro', end = '')
            wait()
            print()
            print('Livro adicionado a nossa base de dados!')
            sleep(1)
            workbook.save(BOOK_PATH)
            
            
            
        case '3':
            removebook = str(input('Qual livro gostaria de remover da nossa biblioteca? '))
            found = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == removebook:
                        found = True
                        idx = row[0].row
                        sure = str(input(f'Tem certeza que deseja remover {removebook} linha: {idx} [S]/[N]')).strip().upper()
                        if sure == 'S':
                            worksheet.delete_rows(idx,1)
                            print('Livro removido com sucesso.')
                            workbook.save(BOOK_PATH)
                        else:
                            print('Livro não removido!')
            if not found:
                print('Livro não encontrado.')
            sleep(3)
            
            
            
        case '4':
            print('Volte sempre!')
            sleep(2)
            quit()
            
        case _:
            print('Opção inválida :(')
            sleep(2)
